import tkinter as tk
from tkinter import *
from tkinter import ttk
from datetime import datetime
import pathlib
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from PIL import ImageTk, Image
import pandas as pd
import numpy as np
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import KFold
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import queue
from tkinter import messagebox
import re


root = tk.Tk()


class App():
    def __init__(self, master):
        self.window = master
        self.window.title("Lung Cancer Diasese")
        self.window.resizable(False, False)
        self.icon = ImageTk.PhotoImage(Image.open(
            r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\lung cancer disease\projek sda sem 2\Picture3.png"))
        self.window.iconphoto(False, self.icon)
        self.font = Font(bold=True)
        self.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thin', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'))

        self.alignment = Alignment(horizontal='center', vertical='center')
        self.file = pathlib.Path(
            r'C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\lung cancer disease\projek sda sem 2\data.xlsx')
        if self.file.exists():
            pass
        else:
            self.file = Workbook()
            sheet = self.file.active
            headers = ['Hari', 'Waktu', 'Nama Pasien', 'Umur', 'Jenis Kelamin',
                       'Nomer Telepon', 'Alamat', 'Jaminan Kesehatan',
                       'Air Pollution', 'Alcohol use', 'Dust Allergy', 'Occupational Hazards',
                       'Genetic Risk', 'Chronic Lung Disease', 'Balanced Diet', 'Obesity', 'Smoking ',
                       'Passive Smoker', 'Chest Pain', 'Coughing of Blood', 'Fatigue', 'Weight Loss',
                       'Shortness of Breath', 'Wheezing', 'Swallowing Difficulty', 'Clubbing of Finger Nails',
                       'Frequent Cold', 'Dry Cough', 'Snoring', 'Level']

            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = self.font
                cell.border = self.border
                cell.alignment = self.alignment
            self.file.save(
                r'C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\lung cancer disease\projek sda sem 2\data.xlsx')

        self.style = ttk.Style(self.window)
        self.window.tk.call(
            "source", r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\lung cancer disease\projek sda sem 2\forest-dark.tcl")
        self.window.tk.call(
            "source", r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\lung cancer disease\projek sda sem 2\forest-light.tcl")
        self.style.theme_use("forest-dark")

        self.combo_list = ["Very Low", "Low", "Below Average", 'Average',
                           'Above Average', 'High', 'Very High', "Maximum", 'Peak']

        self.combo_list8 = ["Very Low", "Low", "Below Average", 'Average',
                            'Above Average', 'High', 'Very High', "Maximum"]

        self.combo_list7 = ["Very Low", "Low", 'Average',
                            'Above Average', 'High', 'Very High', "Maximum"]
        self.d = tk.IntVar(value=2)
        self.h = tk.IntVar(value=2)

        self.notebook = ttk.Notebook(self.window)

        self.tab1 = ttk.Frame(self.notebook)
        self.tab2 = ttk.Frame(self.notebook)

        self.notebook.add(self.tab1, text='Tab 1')
        self.notebook.add(self.tab2, text='Tab 2')

        self.notebook.pack(expand=True, fill='both')

        self.frame = ttk.Frame(self.tab1)
        self.frame.pack()

        self.frame2 = ttk.Frame(self.tab2)
        self.frame2.pack()

        self.frame3 = ttk.Frame(self.tab2)
        self.frame3.pack()

        self.frame4 = ttk.Frame(self.tab2)
        self.frame4.pack()

        self.biodata_row = ttk.LabelFrame(self.frame, text="Biodata Row")
        self.biodata_row.grid(row=0, column=0, padx=20, pady=10)

        self.bpjs_row = ttk.Labelframe(self.biodata_row, text='bpjs')
        self.bpjs_row.grid(row=3, column=0, padx=3, pady=(4))

        radio_1 = ttk.Radiobutton(
            self.bpjs_row, text="NON BPJS", variable=self.d, value=1)
        radio_1.grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
        radio_2 = ttk.Radiobutton(
            self.bpjs_row, text="BPJS", variable=self.d, value=2)
        radio_2.grid(row=1, column=0, padx=5, pady=10, sticky="nsew")

        radio_3 = ttk.Radiobutton(
            self.bpjs_row, text="Male", variable=self.h, value=1)
        radio_3.grid(row=0, column=1, padx=5, pady=10, sticky="nsew")
        radio_4 = ttk.Radiobutton(
            self.bpjs_row, text="Female", variable=self.h, value=2)
        radio_4.grid(row=1, column=1, padx=5, pady=10, sticky="nsew")

        self.biodata_row1 = ttk.LabelFrame(self.frame, text="Keluhan Row")
        self.biodata_row1.grid(row=1, column=0, padx=20, pady=10)

        complaint_row = ttk.Notebook(self.biodata_row1)
        complaint_row.grid(row=1, column=0, padx=20, pady=10)
        test1 = ttk.Frame(complaint_row)
        test2 = ttk.Frame(complaint_row)
        test3 = ttk.Frame(complaint_row)
        complaint_row.add(test1, text='test1')
        complaint_row.add(test2, text='test2')
        complaint_row.add(test3, text='test3')
        self.name_entry = ttk.Entry(self.biodata_row)
        self.name_entry.insert(0, "Name")
        self.name_entry.bind(
            "<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.name_entry.grid(row=0, column=0, padx=10,
                             pady=(20, 20), sticky="ew")
        self.save_button = ttk.Button(test3, text="Save to Predict",
                                      style="Accent.TButton", command=self.knn)
        self.save_button.grid(row=4, column=2, columnspan=2,
                              padx=10, pady=20, sticky="ew")

        label_airpollution = ttk.Label(test1, text='Air Pollution')
        label_airpollution.grid(row=0, column=0, padx=10, pady=10)
        label_alcoholuse = ttk.Label(test1, text='Alcohol use')
        label_alcoholuse.grid(row=0, column=1, padx=10, pady=10)
        label_dustallergy = ttk.Label(test1, text='Dust Allergy')
        label_dustallergy.grid(row=0, column=2, padx=10, pady=10)
        label_occupationalhazards = ttk.Label(
            test1, text='OccuPational Hazards')
        label_occupationalhazards.grid(row=0, column=3, padx=10, pady=10)

        label_geneticrisk = ttk.Label(test1, text='Genetic Risk')
        label_geneticrisk.grid(row=2, column=0, padx=10, pady=10)
        label_chroniclung = ttk.Label(test1, text='Chronic Lung Disease')
        label_chroniclung.grid(row=2, column=1, padx=10, pady=10)
        label_balanceddiet = ttk.Label(test1, text='Balanced Diet')
        label_balanceddiet.grid(row=2, column=2, padx=10, pady=10)
        label_obesity = ttk.Label(test1, text='Obesity')
        label_obesity.grid(row=2, column=3, padx=10, pady=10)

        label_smoking = ttk.Label(test2, text='Smoking')
        label_smoking.grid(row=0, column=0, padx=10, pady=10)
        label_passivesmoker = ttk.Label(test2, text='Passive Smoker')
        label_passivesmoker.grid(row=0, column=1, padx=10, pady=10)
        label_chestpain = ttk.Label(test2, text='Chest Pain')
        label_chestpain.grid(row=0, column=2, padx=10, pady=10)
        label_coughingofblood = ttk.Label(test2, text='Coughing of Blood')
        label_coughingofblood.grid(row=0, column=3, padx=10, pady=10)

        label_fatigue = ttk.Label(test2, text='Fatigue')
        label_fatigue.grid(row=2, column=0, padx=10, pady=10)
        label_weightloss = ttk.Label(test2, text='Weight Loss')
        label_weightloss.grid(row=2, column=1, padx=10, pady=10)
        label_shortnessofbreath = ttk.Label(test2, text='Shortness of Breath')
        label_shortnessofbreath.grid(row=2, column=2, padx=10, pady=10)
        label_wheezing = ttk.Label(test2, text='Wheezing')
        label_wheezing.grid(row=2, column=3, padx=10, pady=10)

        label_swalowingdiffi = ttk.Label(test3, text='Swallowing DIfficulty')
        label_swalowingdiffi.grid(row=0, column=0, padx=10, pady=10)
        label_clubbing = ttk.Label(test3, text='Clubbing of Finger Nails')
        label_clubbing.grid(row=0, column=1, padx=10, pady=10)
        label_frequentcold = ttk.Label(test3, text='Frequent Cold')
        label_frequentcold.grid(row=0, column=2, padx=10, pady=10)
        label_drycough = ttk.Label(test3, text='Dry Cough')
        label_drycough.grid(row=0, column=3, padx=10, pady=10)

        label_snoring = ttk.Label(test3, text='Snoring')
        label_snoring.grid(row=2, column=1, padx=10, pady=10, columnspan=2)

        self.alamat_entry = ttk.Entry(self.biodata_row)
        self.alamat_entry.insert(0, "Address")
        self.alamat_entry.bind(
            "<FocusIn>", lambda e: self.alamat_entry.delete('0', 'end'))
        self.alamat_entry.grid(row=0, column=2, padx=10,
                               pady=(20, 20), sticky="ew")

        self.age_spinbox = ttk.Spinbox(self.biodata_row, from_=18, to=100)
        self.age_spinbox.insert(0, "Age")
        self.age_spinbox.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

        self.phone_entry = ttk.Entry(self.biodata_row)
        self.phone_entry.insert(0, "Phone Number")
        self.phone_entry.bind(
            "<FocusIn>", lambda e: self.phone_entry.delete('0', 'end'))
        self.phone_entry.grid(row=1, column=0, padx=10,
                              pady=(20, 10), sticky="ew")

        ######### segment combo cox#############
        self.airpollution_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list8, width=10)
        self.airpollution_combobox.current(0)
        self.airpollution_combobox.grid(
            row=1, column=0, padx=15, pady=5, sticky="ew")

        self.alcoholuse_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list8, width=10)
        self.alcoholuse_combobox.current(0)
        self.alcoholuse_combobox.grid(
            row=1, column=1, padx=15, pady=5, sticky="ew")

        self.dustalergy_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list8, width=10)
        self.dustalergy_combobox.current(0)
        self.dustalergy_combobox.grid(
            row=1, column=2, padx=15, pady=5, sticky="ew")

        self.occupationalhazard_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list8, width=10)
        self.occupationalhazard_combobox.current(0)
        self.occupationalhazard_combobox.grid(
            row=1, column=3, padx=15, pady=5, sticky="ew")

        self.geneticrisk_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list7, width=10)
        self.geneticrisk_combobox.current(0)
        self.geneticrisk_combobox.grid(
            row=3, column=0, padx=15, pady=5, sticky="ew")

        self.croniclung_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list7, width=10)
        self.croniclung_combobox.current(0)
        self.croniclung_combobox.grid(
            row=3, column=1, padx=15, pady=5, sticky="ew")

        self.balanceddiet_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list7, width=10)
        self.balanceddiet_combobox.current(0)
        self.balanceddiet_combobox.grid(
            row=3, column=2, padx=15, pady=5, sticky="ew")

        self.obesity_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list7, width=10)
        self.obesity_combobox.current(0)
        self.obesity_combobox.grid(
            row=3, column=3, padx=15, pady=5, sticky="ew")

        self.smoking_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list8, width=10)
        self.smoking_combobox.current(0)
        self.smoking_combobox.grid(
            row=1, column=0, padx=15, pady=5, sticky="ew")

        self.passivesmoker_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list8, width=10)
        self.passivesmoker_combobox.current(0)
        self.passivesmoker_combobox.grid(
            row=1, column=1, padx=15, pady=5, sticky="ew")

        self.chestpain_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list, width=10)
        self.chestpain_combobox.current(0)
        self.chestpain_combobox.grid(
            row=1, column=2, padx=15, pady=5, sticky="ew")

        self.coughingofblood_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list, width=10)
        self.coughingofblood_combobox.current(0)
        self.coughingofblood_combobox.grid(
            row=1, column=3, padx=15, pady=5, sticky="ew")

        self.fatigue_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list, width=10)
        self.fatigue_combobox.current(0)
        self.fatigue_combobox.grid(
            row=3, column=0, padx=15, pady=5, sticky="ew")

        self.weightloss_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list8, width=10)
        self.weightloss_combobox.current(0)
        self.weightloss_combobox.grid(
            row=3, column=1, padx=15, pady=5, sticky="ew")

        self.shortnessofbreath_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list, width=10)
        self.shortnessofbreath_combobox.current(0)
        self.shortnessofbreath_combobox.grid(
            row=3, column=2, padx=15, pady=5, sticky="ew")

        self.wheezing_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list8, width=10)
        self.wheezing_combobox.current(0)
        self.wheezing_combobox.grid(
            row=3, column=3, padx=15, pady=5, sticky="ew")

        self.swallowingdiffi_combobox = ttk.Combobox(
            test3, state='readonly', values=self.combo_list8, width=10)
        self.swallowingdiffi_combobox.current(0)
        self.swallowingdiffi_combobox.grid(
            row=1, column=0, padx=15, pady=5, sticky="ew")

        self.clubbing_combobox = ttk.Combobox(
            test3, state='readonly', values=self.combo_list, width=10)
        self.clubbing_combobox.current(0)
        self.clubbing_combobox.grid(
            row=1, column=1, padx=15, pady=5, sticky="ew")

        self.drycough_combobox = ttk.Combobox(
            test3, state='readonly', values=self.combo_list7, width=10)
        self.drycough_combobox.current(0)
        self.drycough_combobox.grid(
            row=1, column=3, padx=15, pady=5, sticky="ew")

        self.frequentcold_combobox = ttk.Combobox(
            test3, state='readonly', values=self.combo_list7, width=10)
        self.frequentcold_combobox.current(0)
        self.frequentcold_combobox.grid(
            row=1, column=2, padx=15, pady=5, sticky="ew")

        self.snoring_combobox = ttk.Combobox(
            test3, state='readonly', values=self.combo_list7, width=10)
        self.snoring_combobox.current(0)
        self.snoring_combobox.grid(row=3, column=1, columnspan=2,
                                   padx=15, pady=5, sticky="ew")

        self.a = tk.BooleanVar()
        checkbutton = ttk.Checkbutton(
            self.biodata_row, text="Terms&Conditions", variable=self.a)
        checkbutton.grid(row=4, column=2, padx=5, pady=5, sticky="nsew")

        button = ttk.Button(self.biodata_row, text="Insert",
                            command=self.on_button_click)
        button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

        separator = ttk.Separator(self.biodata_row)
        separator.grid(row=5, column=0, padx=(20, 10), pady=10, sticky="ew")
        separator1 = ttk.Separator(self.biodata_row)
        separator1.grid(row=5, column=1, padx=(20, 10), pady=10, sticky="ew")
        separator2 = ttk.Separator(self.biodata_row)
        separator2.grid(row=5, column=2, padx=(20, 10), pady=10, sticky="ew")

        self.mode_switch = ttk.Checkbutton(
            self.biodata_row, text="Mode", style="Switch", command=self.toggle_mode)
        self.mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

        self.buttontab2 = ttk.Button(
            self.frame2, text='SHOW GRAFIK', style='Accent.TButton', command=self.grafik)
        self.buttontab2.grid(row=0, column=0)

        if self.mode_switch.instate(["selected"]):
            self.mode = 'default'
            self.bg1 = '#FFF'
            self.bg2 = '#FFF'
        else:
            self.mode = 'dark_background'
            self.bg1 = '#313131'
            self.bg2 = '#313131'

        self.biodata_row3 = ttk.LabelFrame(self.frame, text="Antrian")
        self.biodata_row3.grid(row=0, column=1, padx=20, pady=10)

        treeFrame = ttk.Frame(self.biodata_row3)
        treeFrame.grid(row=0, column=1, pady=10)

        self.biodata_row4 = ttk.LabelFrame(
            self.frame, text="Antrian Sudah dilayani")
        self.biodata_row4.grid(row=1, column=1, padx=20, pady=10)

        treeFrame2 = ttk.Frame(self.biodata_row4)
        treeFrame2.grid(row=1, column=1, pady=10)

        treeScroll = ttk.Scrollbar(treeFrame)
        treeScroll.pack(side="right", fill="y")

        treeScroll2 = ttk.Scrollbar(treeFrame2)
        treeScroll2.pack(side="right", fill="y")

        cols = ("Time", 'Day', "Name", "Age", "Gender", "Level")
        self.treeview = ttk.Treeview(treeFrame, show="headings",
                                     yscrollcommand=treeScroll.set, columns=cols, height=13)
        self.treeview2 = ttk.Treeview(treeFrame2, show="headings",
                                      yscrollcommand=treeScroll.set, columns=cols, height=13)

        self.treeview.heading('#1', text='Level')
        self.treeview.heading('#2', text='Day')
        self.treeview.heading('#3', text='Time')
        self.treeview.heading('#4', text='Name')
        self.treeview.heading('#5', text='Age')
        self.treeview.heading('#6', text='Gender')
        self.treeview.column('Level', width=50)
        self.treeview.column('Day', width=50)
        self.treeview.column('Time', width=70)
        self.treeview.column('Name', width=50)
        self.treeview.column('Age', width=50)
        self.treeview.column('Gender', width=50)

        self.treeview2.heading('#1', text='Level')
        self.treeview2.heading('#2', text='Day')
        self.treeview2.heading('#3', text='Time')
        self.treeview2.heading('#4', text='Name')
        self.treeview2.heading('#5', text='Age')
        self.treeview2.heading('#6', text='Gender')
        self.treeview2.column('Level', width=50)
        self.treeview2.column('Day', width=50)
        self.treeview2.column('Time', width=50)
        self.treeview2.column('Name', width=50)
        self.treeview2.column('Age', width=50)
        self.treeview2.column('Gender', width=50)

        self.treeview.bind("<Double-1>", self.on_click)
        self.treeview.pack()
        treeScroll.config(command=self.treeview.yview)

        self.treeview2.bind("<Double-1>")
        self.treeview2.pack()
        treeScroll2.config(command=self.treeview.yview)

        self.priority_queue()

        self.window.update()
        self.window.minsize(self.window.winfo_width(),
                            self.window.winfo_height())
        x_cordinate = int((self.window.winfo_screenwidth()/2) -
                          (self.window.winfo_width()/2))
        y_cordinate = int((self.window.winfo_screenheight()/2) -
                          (self.window.winfo_height()/2))
        self.window.geometry("+{}+{}".format(x_cordinate, y_cordinate))

    def update_chart(self):
        # hapus chart lama
        self.ax1.clear()
        self.ax2.clear()

    # buat chart baru dengan data terbaru
        plt.style.use(self.mode)
        df = pd.read_excel('data.xlsx')
        nilai_counts = df['Level'].value_counts()
        gender_counts = df['Jenis Kelamin'].value_counts()
        labels = gender_counts.index
        sizes = gender_counts.values
        plt.rcParams["axes.prop_cycle"] = plt.cycler(
            color=['#9CA777', '#64C2A6', '#AADEA7'])
        self.total_categori = nilai_counts
        self.ax2.pie(sizes, labels=labels, autopct='%1.1f%%',
                     startangle=90)
        self.ax2.set_title("Berdasarkan \nJenis Kelamin")
        self.ax1.bar(
            self.total_categori.keys(), self.total_categori)

    # refresh tampilan chart
        self.fig1.patch.set_facecolor(self.bg1)
        self.ax1.patch.set_facecolor(self.bg2)

        self.fig2.patch.set_facecolor(self.bg1)
        self.ax2.patch.set_facecolor(self.bg2)
        self.canvas1.draw()
        self.canvas2.draw()

    def on_button_click(self):
        answer = messagebox.askquestion(
            "Konfirmasi", "Konfirmasi untuk melanjutkan?")
        if answer == 'yes':
            self.insert_row()

    def insert_row(self):
        if not self.name_entry.get() or self.name_entry.get() == 'Name':
            messagebox.showerror("Warning", "Nama tidak boleh kosong")
            return
        elif not self.age_spinbox.get().isdigit():
            messagebox.showerror("Warning", "Umur harus berupa bilangan bulat")
            return
        elif not self.alamat_entry.get() or self.alamat_entry.get() == 'Address':
            messagebox.showerror("Warning", "Alamat harus diisi")
            return
        elif not self.phone_entry.get().isdigit():
            messagebox.showerror("Warning", "Nomer harus diisi")
            return
        elif not self.a.get():
            messagebox.showerror(
                "Warning", "Syarat dan ketentuan harus dicentang")
            return
        elif self.hasiltrain == '':
            messagebox.showerror(
                "Warning", "Prediksi keluhan terlebih dahulu di kolom dibawah!!")
            return
        self.now = datetime.now()
        self.tanggal = self.now.strftime("%d %B %Y")
        self.time_string = self.now.strftime('%H:%M:%S')
        self.name = self.name_entry.get()
        self.age = int(self.age_spinbox.get())
        self.addres = self.alamat_entry.get()
        self.phone = self.phone_entry.get()
        self.bpjs = self.d.get()
        train = self.hasiltrain

        if int(self.d.get()) == 1:
            self.bpjs = "Non BPJS"
        else:
            self.bpjs = "BPJS"

        if int(self.h.get()) == 1:
            self.gender = "Male"
        else:
            self.gender = "Female"

    # Insert row into Excel sheet
        path = self.file
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_values = [self.tanggal, self.time_string, self.name,
                      self.age, self.gender, self.phone, self.addres, self.bpjs, self.airpollution_combobox.get(), self.alcoholuse_combobox.get(), self.dustalergy_combobox.get(), self.occupationalhazard_combobox.get(
                      ), self.geneticrisk_combobox.get(), self.croniclung_combobox.get(), self.balanceddiet_combobox.get(), self.obesity_combobox.get(), self.smoking_combobox.get(), self.passivesmoker_combobox.get(),
                      self.chestpain_combobox.get(), self.coughingofblood_combobox.get(), self.fatigue_combobox.get(), self.weightloss_combobox.get(), self.shortnessofbreath_combobox.get(), self.wheezing_combobox.get(), self.swallowingdiffi_combobox.get(), self.clubbing_combobox.get(), self.frequentcold_combobox.get(), self.drycough_combobox.get(), self.snoring_combobox.get(), train]
        row_values2 = [self.tanggal, self.name,
                       self.age, self.gender, train]
        sheet.append(row_values)
        workbook.save(path)

    # Insert row into treeview
        self.treeview.insert('', tk.END, values=row_values2)

    # Clear the values
        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, "Name")
        self.age_spinbox.delete(0, "end")
        self.age_spinbox.insert(0, "Age")
        self.alamat_entry.delete(0, 'end')
        self.alamat_entry.insert(0, 'Address')
        self.phone_entry.delete(0, 'end')
        self.phone_entry.insert(0, 'Phone Number')
        self.h.set(2)
        self.d.set(2)

        self.airpollution_combobox.set(self.combo_list[0])
        self.alcoholuse_combobox.set(self.combo_list[0])
        self.dustalergy_combobox.set(self.combo_list[0])
        self.occupationalhazard_combobox.set(self.combo_list[0])
        self.geneticrisk_combobox.set(self.combo_list[0])
        self.croniclung_combobox.set(self.combo_list[0])
        self.obesity_combobox.set(self.combo_list[0])
        self.smoking_combobox.set(self.combo_list[0])
        self.passivesmoker_combobox.set(self.combo_list[0])
        self.chestpain_combobox.set(self.combo_list[0])
        self.coughingofblood_combobox.set(self.combo_list[0])
        self.fatigue_combobox.set(self.combo_list[0])
        self.weightloss_combobox.set(self.combo_list[0])
        self.shortnessofbreath_combobox.set(self.combo_list[0])
        self.wheezing_combobox.set(self.combo_list[0])
        self.clubbing_combobox.set(self.combo_list[0])
        self.frequentcold_combobox.set(self.combo_list[0])
        self.drycough_combobox.set(self.combo_list[0])
        self.snoring_combobox.set(self.combo_list[0])
        self.balanceddiet_combobox.set(self.combo_list[0])
        self.swallowingdiffi_combobox.set(self.combo_list[0])

        self.priority_queue()
        self.update_chart()

    def grafik(self):
        dfgraf = pd.read_excel('data.xlsx')
        nilai_counts = dfgraf['Level'].value_counts()
        gender_counts = dfgraf['Jenis Kelamin'].value_counts()
        plt.style.use(self.mode)
        labels = gender_counts.index
        sizes = gender_counts.values
        plt.rcParams["axes.prop_cycle"] = plt.cycler(
            color=['#9CA777', '#64C2A6', '#AADEA7'])
        # colors = ['#64C2A6', '#AADEA7']  # Assigning colors to each category
        self.fig2, self.ax2 = plt.subplots()
        self.ax2.pie(sizes, labels=labels, autopct='%1.1f%%',
                     startangle=90)
        self.ax2.axis('equal')
        self.ax2.set_title("Porsi Persentase \nBerdasarkan jenis kelamin")

        self.total_categori = nilai_counts
        self.fig1, self.ax1 = plt.subplots()
        self.barlist = self.ax1.bar(
            self.total_categori.keys(), self.total_categori)
        self.ax1.set_title("Grafik")
        self.ax1.set_xlabel("Level")
        self.ax1.set_ylabel("Jumlah Pasien")

        self.fig1.patch.set_facecolor(self.bg1)
        self.ax1.patch.set_facecolor(self.bg2)

        self.fig2.patch.set_facecolor(self.bg1)
        self.ax2.patch.set_facecolor(self.bg2)

        self.canvas1 = FigureCanvasTkAgg(self.fig1, self.frame2)
        self.canvas1.draw()
        self.canvas1.get_tk_widget().grid(row=1, column=0)

        self.canvas2 = FigureCanvasTkAgg(self.fig2, self.frame2)
        self.canvas2.draw()
        self.canvas2.get_tk_widget().grid(row=1, column=1)
        self.buttontab2.destroy()
        self.window.update()
        self.window.minsize(self.window.winfo_width(),
                            self.window.winfo_height())
        x_cordinate = int((self.window.winfo_screenwidth()/2) -
                          (self.window.winfo_width()/2))
        y_cordinate = int((self.window.winfo_screenheight()/2) -
                          (self.window.winfo_height()/2))
        self.window.geometry("+{}+{}".format(x_cordinate, y_cordinate))

    def knn(self):
        if not self.name_entry.get() or self.name_entry.get() == 'Name':
            messagebox.showwarning("Warning", "isi lengkap Biodata row")
            return
        elif not self.age_spinbox.get().isdigit():
            messagebox.showwarning("Warning", "isi lengkap Biodata row")
            return
        elif not self.alamat_entry.get() or self.alamat_entry.get() == 'Address':
            messagebox.showwarning("Warning", "isi lengkap Biodata row")
            return
        elif not self.phone_entry.get().isdigit():
            messagebox.showwarning("Warning", "isi lengkap Biodata row")
            return
        elif not self.a.get():
            messagebox.showwarning(
                "Warning", "isi lengkap Biodata row")
            return

        age = int(self.age_spinbox.get())
        gender = int(self.h.get())
        s = pd.read_csv('cancer patient data sets.csv')
        s = s.values
        x = s[:, 2:25]
        y = s[:, 25]
        kf = KFold(n_splits=10, random_state=0, shuffle=True)
        variables9 = [self.airpollution_combobox, self.alcoholuse_combobox, self.dustalergy_combobox, self.occupationalhazard_combobox, self.geneticrisk_combobox, self.croniclung_combobox, self.balanceddiet_combobox, self.obesity_combobox, self.smoking_combobox, self.passivesmoker_combobox,
                      self.chestpain_combobox, self.coughingofblood_combobox, self.fatigue_combobox, self.weightloss_combobox, self.shortnessofbreath_combobox, self.wheezing_combobox, self.swallowingdiffi_combobox, self.clubbing_combobox, self.drycough_combobox, self.frequentcold_combobox, self.snoring_combobox]
        variables8 = [self.geneticrisk_combobox, self.croniclung_combobox, self.balanceddiet_combobox,
                      self.obesity_combobox, self.frequentcold_combobox, self.drycough_combobox, self.snoring_combobox]
        for variable in variables9:
            if variable.get() == 'Very Low':
                variable.set('1')
                variable.set(int(variable.get()))
            elif variable.get() == 'Low':
                variable.set('2')
                variable.set(int(variable.get()))
            elif variable == variables8 and variable.get() == 'Average':
                variable.set('3')
                variable.set(int(variable.get()))
            elif variable.get() == 'Below Average':
                variable.set('3')
                variable.set(int(variable.get()))
            elif variable.get() == 'Average':
                variable.set('4')
                variable.set(int(variable.get()))
            elif variable.get() == 'Above Average':
                variable.set('5')
                variable.set(int(variable.get()))
            elif variable.get() == 'High':
                variable.set('6')
                variable.set(int(variable.get()))
            elif variable.get() == 'Very High':
                variable.set('7')
                variable.set(int(variable.get()))
            elif variable.get() == 'Maximum':
                variable.set('8')
                variable.set(int(variable.get()))
            elif variable.get() == 'Peak':
                variable.set('9')
                variable.set(int(variable.get()))

        neigh2 = KNeighborsClassifier(n_neighbors=2, metric='euclidean')
        neigh2.fit(x, y)

        train = neigh2.predict(
            [[age, gender,
              int(self.airpollution_combobox.get()),
              int(self.alcoholuse_combobox.get()),
              int(self.dustalergy_combobox.get()),
              int(self.occupationalhazard_combobox.get()),
              int(self.geneticrisk_combobox.get()),
              int(self.croniclung_combobox.get()),
              int(self.balanceddiet_combobox.get()),
              int(self.obesity_combobox.get()),
              int(self.smoking_combobox.get()),
              int(self.passivesmoker_combobox.get()),
              int(self.chestpain_combobox.get()),
              int(self.coughingofblood_combobox.get()),
              int(self.fatigue_combobox.get()),
              int(self.weightloss_combobox.get()),
              int(self.shortnessofbreath_combobox.get()),
              int(self.wheezing_combobox.get()),
              int(self.swallowingdiffi_combobox.get()),
              int(self.clubbing_combobox.get()),
              int(self.drycough_combobox.get()),
              int(self.frequentcold_combobox.get()),
              int(self.snoring_combobox.get())]])

        self.hasiltrain = train[0]
        messagebox.showinfo("Prediksi Level cancer", self.hasiltrain)

    def toggle_mode(self):
        if self.mode_switch.instate(["selected"]):
            self.style.theme_use("forest-light")
            self.mode_switch.configure(text="Light Mode")
            self.mode = 'default'
            self.bg1 = '#ffffff'
            self.bg2 = '#ffffff'
            self.update_chart()

        else:
            self.style.theme_use("forest-dark")
            self.mode_switch.configure(text="Dark Mode")
            self.mode = 'dark_background'
            self.bg1 = '#313131'
            self.bg2 = '#313131'
            self.update_chart()

    def on_click(self, event):
        answer = messagebox.askquestion(
            "Konfirmasi", "Apakah Sudah dilayani")
        if answer == 'yes':
            selected_item = self.treeview.selection()[0]  # get selected item
            # get values from selected item
            values = self.treeview.item(selected_item)['values']
            # insert values into other treeview
            self.treeview2.insert('', 'end', values=values)
            # delete selected item from original treeview
            self.treeview.delete(selected_item)

    def priority_queue(self):
        df = pd.read_excel("data.xlsx")
        self.order = {'Low': 2, 'Medium': 1, 'High': 0}

        # Bubble sort
        n = len(df)
        for i in range(n):
            for j in range(0, n-i-1):
                if (self.order[df.iloc[j]['Level']] > self.order[df.iloc[j+1]['Level']]) or (self.order[df.iloc[j]['Level']] == self.order[df.iloc[j+1]['Level']] and df.iloc[j]['Hari'] > df.iloc[j+1]['Hari']) or (self.order[df.iloc[j]['Level']] == self.order[df.iloc[j+1]['Level']] and df.iloc[j]['Hari'] == df.iloc[j+1]['Hari'] and df.iloc[j]['Waktu'] > df.iloc[j+1]['Waktu']):
                    df.iloc[j], df.iloc[j+1] = df.iloc[j+1], df.iloc[j]

        q = queue.PriorityQueue()
        for index, row in df.iterrows():
            q.put((self.order[row['Level']], row['Hari'], row['Waktu'],
                  row['Nama Pasien'], row['Umur'], row['Jenis Kelamin']))

        for i in self.treeview.get_children():
            self.treeview.delete(i)

        order_dict = {v: k for k, v in self.order.items()}
        for item in sorted(q.queue):
            item = list(item)
            item[0] = order_dict[item[0]]
            self.treeview.insert("", "end", values=tuple(item))


app = App(root)
root.mainloop()

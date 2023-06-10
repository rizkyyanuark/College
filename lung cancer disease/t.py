import tkinter as tk
from tkinter import *
from tkinter import ttk
from datetime import datetime
import pathlib
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from PIL import ImageTk, Image
import pandas as pd
import numpy as np
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import KFold
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import recall_score
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


root = tk.Tk()


class App():
    def __init__(self, master):
        self.window = master
        self.window.title("Lung Cancer Diasese")
        # root.geometry("950x700")
        self.window.resizable(False, False)
        self.icon = ImageTk.PhotoImage(Image.open("Picture3.png"))
        self.window.iconphoto(False, self.icon)
        self.font = Font(bold=True)
        self.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thin', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'))

        self.alignment = Alignment(horizontal='center', vertical='center')
        self.file = pathlib.Path('data1 copy.xlsx')
        if self.file.exists():
            pass
        else:
            self.file = Workbook()
            sheet = self.file.active
            headers = ['Nama Pasien', 'Umur', 'Jenis Kelamin',
                       'Nomer Telepon', 'Alamat', 'Nomor Telepon' 'Level']
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = self.font
                cell.border = self.border
                cell.alignment = self.alignment
            self.file.save('data1 copy.xlsx')
        root.columnconfigure(index=0, weight=1)
        root.columnconfigure(index=1, weight=1)
        root.rowconfigure(index=0, weight=1)
        root.rowconfigure(index=1, weight=1)

        # self.style = ttk.Style(self.window)
        # self.window.tk.call("source", "forest-dark.tcl")
        # self.window.tk.call("source", "forest-light.tcl")
        # self.style.theme_use("forest-dark")

        self.combo_list = ["Very Low", "Low", "Below Average", 'Average',
                           'Above Average', 'High', 'Very High', "Maximum", 'Peak']

        self.combo_list8 = ["Very Low", "Low", "Below Average", 'Average',
                            'Above Average', 'High', 'Very High', "Maximum"]

        self.combo_list7 = ["Very Low", "Low", 'Average',
                            'Above Average', 'High', 'Very High', "Maximum"]

        self.d = tk.IntVar(value=2)
        self.h = tk.IntVar(value=2)

        self.notebook = ttk.Notebook(self.window)

        self.tab1 = ttk.Frame(self.notebook)
        self.tab2 = ttk.Frame(self.notebook)

        self.notebook.add(self.tab1, text='Tab 1')
        self.notebook.add(self.tab2, text='Tab 2')

        self.notebook.pack(expand=1, fill='both')

        self.frame = ttk.Frame(self.tab1)
        self.frame.pack()

        self.biodata_row = ttk.LabelFrame(self.frame, text="Biodata Row")
        self.biodata_row.grid(row=0, column=0, padx=20, pady=10)

        self.bpjs_row = ttk.Labelframe(self.biodata_row, text='bpjs')
        self.bpjs_row.grid(row=3, column=0, padx=3, pady=(4))

        radio_1 = ttk.Radiobutton(
            self.bpjs_row, text="NON BPJS", variable=self.d, value=1)
        radio_1.grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
        radio_2 = ttk.Radiobutton(
            self.bpjs_row, text="BPJS", variable=self.d, value=2)
        radio_2.grid(row=1, column=0, padx=5, pady=10, sticky="nsew")

        radio_3 = ttk.Radiobutton(
            self.bpjs_row, text="Male", variable=self.h, value=1)
        radio_3.grid(row=0, column=1, padx=5, pady=10, sticky="nsew")
        radio_4 = ttk.Radiobutton(
            self.bpjs_row, text="Female", variable=self.h, value=2)
        radio_4.grid(row=1, column=1, padx=5, pady=10, sticky="nsew")

        complaint_row = ttk.Notebook(self.frame)
        complaint_row.grid(row=1, column=0, padx=20, pady=10)
        test1 = ttk.Frame(complaint_row)
        test2 = ttk.Frame(complaint_row)
        test3 = ttk.Frame(complaint_row)
        complaint_row.add(test1, text='test1')
        complaint_row.add(test2, text='test2')
        complaint_row.add(test3, text='test3')

        self.name_entry = ttk.Entry(self.biodata_row)
        self.name_entry.insert(0, "Name")
        self.name_entry.bind(
            "<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.name_entry.grid(row=0, column=0, padx=10,
                             pady=(20, 20), sticky="ew")

        self.save_button = ttk.Button(test3, text="Save to Train",
                                      style="Accent.TButton", command=self.knn)
        self.save_button.grid(row=4, column=2, columnspan=2,
                              padx=10, pady=20, sticky="ew")

        label_airpollution = ttk.Label(test1, text='Air Pollution')
        label_airpollution.grid(row=0, column=0, padx=10, pady=10)
        label_alcoholuse = ttk.Label(test1, text='Alcohol use')
        label_alcoholuse.grid(row=0, column=1, padx=10, pady=10)
        label_dustallergy = ttk.Label(test1, text='Dust Allergy')
        label_dustallergy.grid(row=0, column=2, padx=10, pady=10)
        label_occupationalhazards = ttk.Label(
            test1, text='OccuPational Hazards')
        label_occupationalhazards.grid(row=0, column=3, padx=10, pady=10)

        label_geneticrisk = ttk.Label(test1, text='Genetic Risk')
        label_geneticrisk.grid(row=2, column=0, padx=10, pady=10)
        label_chroniclung = ttk.Label(test1, text='Chronic Lung Disease')
        label_chroniclung.grid(row=2, column=1, padx=10, pady=10)
        label_balanceddiet = ttk.Label(test1, text='Balanced Diet')
        label_balanceddiet.grid(row=2, column=2, padx=10, pady=10)
        label_obesity = ttk.Label(test1, text='Obesity')
        label_obesity.grid(row=2, column=3, padx=10, pady=10)

        label_smoking = ttk.Label(test2, text='Smoking')
        label_smoking.grid(row=0, column=0, padx=10, pady=10)
        label_passivesmoker = ttk.Label(test2, text='Passive Smoker')
        label_passivesmoker.grid(row=0, column=1, padx=10, pady=10)
        label_chestpain = ttk.Label(test2, text='Chest Pain')
        label_chestpain.grid(row=0, column=2, padx=10, pady=10)
        label_coughingofblood = ttk.Label(test2, text='Coughing of Blood')
        label_coughingofblood.grid(row=0, column=3, padx=10, pady=10)

        label_fatigue = ttk.Label(test2, text='Fatigue')
        label_fatigue.grid(row=2, column=0, padx=10, pady=10)
        label_weightloss = ttk.Label(test2, text='Weight Loss')
        label_weightloss.grid(row=2, column=1, padx=10, pady=10)
        label_shortnessofbreath = ttk.Label(test2, text='Shortness of Breath')
        label_shortnessofbreath.grid(row=2, column=2, padx=10, pady=10)
        label_wheezing = ttk.Label(test2, text='Wheezing')
        label_wheezing.grid(row=2, column=3, padx=10, pady=10)

        label_swalowingdiffi = ttk.Label(test3, text='Swallowing DIfficulty')
        label_swalowingdiffi.grid(row=0, column=0, padx=10, pady=10)
        label_clubbing = ttk.Label(test3, text='Clubbing of Finger Nails')
        label_clubbing.grid(row=0, column=1, padx=10, pady=10)
        label_frequentcold = ttk.Label(test3, text='Frequent Cold')
        label_frequentcold.grid(row=0, column=2, padx=10, pady=10)
        label_drycough = ttk.Label(test3, text='Dry Cough')
        label_drycough.grid(row=0, column=3, padx=10, pady=10)

        label_snoring = ttk.Label(test3, text='Snoring')
        label_snoring.grid(row=2, column=1, padx=10, pady=10, columnspan=2)

        self.alamat_entry = ttk.Entry(self.biodata_row)
        self.alamat_entry.insert(0, "Address")
        self.alamat_entry.bind(
            "<FocusIn>", lambda e: self.alamat_entry.delete('0', 'end'))
        self.alamat_entry.grid(row=0, column=2, padx=10,
                               pady=(20, 20), sticky="ew")

        self.age_spinbox = ttk.Spinbox(self.biodata_row, from_=18, to=100)
        self.age_spinbox.insert(0, "Age")
        self.age_spinbox.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

        self.phone_entry = ttk.Entry(self.biodata_row)
        self.phone_entry.insert(0, "Phone Number")
        self.phone_entry.bind(
            "<FocusIn>", lambda e: self.phone_entry.delete('0', 'end'))
        self.phone_entry.grid(row=1, column=0, padx=10,
                              pady=(20, 10), sticky="ew")

        ######### segment combo cox#############
        self.airpollution_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list8, width=10)
        self.airpollution_combobox.current(0)
        self.airpollution_combobox.grid(
            row=1, column=0, padx=15, pady=5, sticky="ew")

        self.alcoholuse_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list8, width=10)
        self.alcoholuse_combobox.current(0)
        self.alcoholuse_combobox.grid(
            row=1, column=1, padx=15, pady=5, sticky="ew")

        self.dustalergy_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list8, width=10)
        self.dustalergy_combobox.current(0)
        self.dustalergy_combobox.grid(
            row=1, column=2, padx=15, pady=5, sticky="ew")

        self.occupationalhazard_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list8, width=10)
        self.occupationalhazard_combobox.current(0)
        self.occupationalhazard_combobox.grid(
            row=1, column=3, padx=15, pady=5, sticky="ew")

        self.geneticrisk_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list7, width=10)
        self.geneticrisk_combobox.current(0)
        self.geneticrisk_combobox.grid(
            row=3, column=0, padx=15, pady=5, sticky="ew")

        self.croniclung_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list7, width=10)
        self.croniclung_combobox.current(0)
        self.croniclung_combobox.grid(
            row=3, column=1, padx=15, pady=5, sticky="ew")

        self.balanceddiet_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list7, width=10)
        self.balanceddiet_combobox.current(0)
        self.balanceddiet_combobox.grid(
            row=3, column=2, padx=15, pady=5, sticky="ew")

        self.obesity_combobox = ttk.Combobox(
            test1, state='readonly', values=self.combo_list7, width=10)
        self.obesity_combobox.current(0)
        self.obesity_combobox.grid(
            row=3, column=3, padx=15, pady=5, sticky="ew")

        self.smoking_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list8, width=10)
        self.smoking_combobox.current(0)
        self.smoking_combobox.grid(
            row=1, column=0, padx=15, pady=5, sticky="ew")

        self.passivesmoker_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list8, width=10)
        self.passivesmoker_combobox.current(0)
        self.passivesmoker_combobox.grid(
            row=1, column=1, padx=15, pady=5, sticky="ew")

        self.chestpain_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list, width=10)
        self.chestpain_combobox.current(0)
        self.chestpain_combobox.grid(
            row=1, column=2, padx=15, pady=5, sticky="ew")

        self.coughingofblood_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list, width=10)
        self.coughingofblood_combobox.current(0)
        self.coughingofblood_combobox.grid(
            row=1, column=3, padx=15, pady=5, sticky="ew")

        self.fatigue_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list, width=10)
        self.fatigue_combobox.current(0)
        self.fatigue_combobox.grid(
            row=3, column=0, padx=15, pady=5, sticky="ew")

        self.weightloss_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list8, width=10)
        self.weightloss_combobox.current(0)
        self.weightloss_combobox.grid(
            row=3, column=1, padx=15, pady=5, sticky="ew")

        self.shortnessofbreath_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list, width=10)
        self.shortnessofbreath_combobox.current(0)
        self.shortnessofbreath_combobox.grid(
            row=3, column=2, padx=15, pady=5, sticky="ew")

        self.wheezing_combobox = ttk.Combobox(
            test2, state='readonly', values=self.combo_list8, width=10)
        self.wheezing_combobox.current(0)
        self.wheezing_combobox.grid(
            row=3, column=3, padx=15, pady=5, sticky="ew")

        self.swallowingdiffi_combobox = ttk.Combobox(
            test3, state='readonly', values=self.combo_list8, width=10)
        self.swallowingdiffi_combobox.current(0)
        self.swallowingdiffi_combobox.grid(
            row=1, column=0, padx=15, pady=5, sticky="ew")

        self.clubbing_combobox = ttk.Combobox(
            test3, state='readonly', values=self.combo_list, width=10)
        self.clubbing_combobox.current(0)
        self.clubbing_combobox.grid(
            row=1, column=1, padx=15, pady=5, sticky="ew")

        self.drycough_combobox = ttk.Combobox(
            test3, state='readonly', values=self.combo_list7, width=10)
        self.drycough_combobox.current(0)
        self.drycough_combobox.grid(
            row=1, column=3, padx=15, pady=5, sticky="ew")

        self.frequentcold_combobox = ttk.Combobox(
            test3, state='readonly', values=self.combo_list7, width=10)
        self.frequentcold_combobox.current(0)
        self.frequentcold_combobox.grid(
            row=1, column=2, padx=15, pady=5, sticky="ew")

        self.snoring_combobox = ttk.Combobox(
            test3, state='readonly', values=self.combo_list7, width=10)
        self.snoring_combobox.current(0)
        self.snoring_combobox.grid(row=3, column=1, columnspan=2,
                                   padx=15, pady=5, sticky="ew")
        progress = ttk.Progressbar(self.tab1, orient="horizontal",
                                   length=200, mode="determinate")
        progress.pack(pady=50)

        # def update_progress():
        #     value = name_entry.get()  # Mendapatkan nilai dari entry
        #     # Mengatur nilai progress bar berdasarkan panjang string
        #     progress['value'] = len(value)
        #     progress.update()  # Memperbarui tampilan progress bar

        # def start():
        #     progress["value"] = 0
        #     progress["maximum"] = 100
        #     while progress["value"] < progress["maximum"]:
        #         progress["value"] += 1

        # button = tk.Button(tab1, text="Start", command=bar)
        # button.pack()

        self.a = tk.BooleanVar()
        checkbutton = ttk.Checkbutton(
            self.biodata_row, text="Terms&Conditions", variable=self.a)
        checkbutton.grid(row=4, column=2, padx=5, pady=5, sticky="nsew")

        # checkbutton = ttk.Checkbutton(self.biodata_row, text="Employed", variable=a)
        # checkbutton.grid(row=7, column=2, padx=10, pady=10, sticky="nsew")

        button = ttk.Button(self.biodata_row, text="Insert",
                            command=self.insert_row)
        button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

        separator = ttk.Separator(self.biodata_row)
        separator.grid(row=5, column=0, padx=(20, 10), pady=10, sticky="ew")
        separator1 = ttk.Separator(self.biodata_row)
        separator1.grid(row=5, column=1, padx=(20, 10), pady=10, sticky="ew")
        separator2 = ttk.Separator(self.biodata_row)
        separator2.grid(row=5, column=2, padx=(20, 10), pady=10, sticky="ew")

        # mode_switch = ttk.Checkbutton(
        #     self.biodata_row, text="Mode", style="Switch", command=self.toggle_mode)
        # mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

        treeFrame = ttk.Frame(self.frame)
        treeFrame.grid(row=0, column=1, pady=10)

        treeScroll = ttk.Scrollbar(treeFrame)
        treeScroll.pack(side="right", fill="y")

        cols = ("Nama Pasien", "Umur", "Jenis Kelamin",
                "Alamat", "Nomor Telepon", "Level")
        self.treeview = ttk.Treeview(treeFrame, show="headings",
                                     yscrollcommand=treeScroll.set, columns=cols, height=13)
        self.treeview.heading('#1', text='Nama Paien')
        self.treeview.heading('#2', text='Umur')
        self.treeview.heading('#3', text='Jenis Kelamin')
        self.treeview.heading('#4', text='Alamat')
        self.treeview.heading('#5', text='Nomor Telepon')
        self.treeview.heading('#6', text='Level')
        self.treeview.column('Nama Pasien', width=50)
        self.treeview.column('Umur', width=50)
        self.treeview.column('Jenis Kelamin', width=50)
        self.treeview.column('Alamat', width=50)
        self.treeview.column('Nomor Telepon', width=50)
        self.treeview.column('Level', width=50)

        self.treeview.bind("<Double-1>", self.on_click)
        self.treeview.pack()
        treeScroll.config(command=self.treeview.yview)

        self.load_data()
        self.grafik()

        root.update()
        root.minsize(root.winfo_width(), root.winfo_height())
        x_cordinate = int((root.winfo_screenwidth()/2) -
                          (root.winfo_width()/2))
        y_cordinate = int((root.winfo_screenheight()/2) -
                          (root.winfo_height()/2))
        root.geometry("+{}+{}".format(x_cordinate, y_cordinate))\


    def load_data(self):
        # Read the Excel file and create a DataFrame
        self.df = pd.read_excel("data1 copy.xlsx")
        for index, row in self.df.iterrows():
            self.treeview.insert("", "end", values=(
                row['Nama Pasien'], row["Umur"], row['Jenis Kelamin'], row['Alamat'], row['Nomor Telepon'], row['Level']))

    def insert_row(self):
        self.now = datetime.now()
        self.tanggal = self.now.strftime("%d:%m:%Y")
        self.name = self.name_entry.get()
        self.age = int(self.age_spinbox.get())
        self.addres = self.alamat_entry.get()
        self.phone = self.phone_entry.get()

        if int(self.h.get()) == 1:
            self.gender = "male"
        else:
            self.gender = "female"

    # Insert row into Excel sheet
        path = self.file
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_values = [self.tanggal, self.name,
                      self.age, self.gender, self.phone, self.addres, self.hasiltrain]
        row_values2 = [self.tanggal, self.name,
                       self.age, self.gender, self.hasiltrain]
        sheet.append(row_values)
        workbook.save(path)

# Insert row into treeview
        self.treeview.insert('', tk.END, values=row_values2)

# Clear the values
        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, "Name")
        self.age_spinbox.delete(0, "end")
        self.age_spinbox.insert(0, "Age")
        self.alamat_entry.delete(0, 'end')
        self.alamat_entry.insert(0, 'Address')
        self.phone_entry.delete(0, 'end')
        self.phone_entry.insert(0, 'Phone Number')

        self.airpollution_combobox.set(self.combo_list[0])
        self.alcoholuse_combobox.set(self.combo_list[0])
        self.dustalergy_combobox.set(self.combo_list[0])
        self.occupationalhazard_combobox.set(self.combo_list[0])
        self.geneticrisk_combobox.set(self.combo_list[0])
        self.croniclung_combobox.set(self.combo_list[0])
        self.obesity_combobox.set(self.combo_list[0])
        self.smoking_combobox.set(self.combo_list[0])
        self.passivesmoker_combobox.set(self.combo_list[0])
        self.chestpain_combobox.set(self.combo_list[0])
        self.coughingofblood_combobox.set(self.combo_list[0])
        self.fatigue_combobox.set(self.combo_list[0])
        self.weightloss_combobox.set(self.combo_list[0])
        self.shortnessofbreath_combobox.set(self.combo_list[0])
        self.wheezing_combobox.set(self.combo_list[0])
        self.clubbing_combobox.set(self.combo_list[0])
        self.frequentcold_combobox.set(self.combo_list[0])
        self.drycough_combobox.set(self.combo_list[0])
        self.snoring_combobox.set(self.combo_list[0])

    def knn(self):
        age = int(self.age_spinbox.get())
        gender = int(self.h.get())
        s = pd.read_csv('cancer patient data sets (3).csv')
        s = s.values
        x = s[:, 2:25]
        y = s[:, 25]
        kf = KFold(n_splits=10, random_state=0, shuffle=True)
        variables9 = [self.airpollution_combobox, self.alcoholuse_combobox, self.dustalergy_combobox, self.occupationalhazard_combobox, self.geneticrisk_combobox, self.croniclung_combobox, self.balanceddiet_combobox, self.obesity_combobox, self.smoking_combobox, self.passivesmoker_combobox,
                      self.chestpain_combobox, self.coughingofblood_combobox, self.fatigue_combobox, self.weightloss_combobox, self.shortnessofbreath_combobox, self.wheezing_combobox, self.swallowingdiffi_combobox, self.clubbing_combobox, self.drycough_combobox, self.frequentcold_combobox, self.snoring_combobox]
        variables8 = [self.geneticrisk_combobox, self.croniclung_combobox, self.balanceddiet_combobox,
                      self.obesity_combobox, self.frequentcold_combobox, self.drycough_combobox, self.snoring_combobox]
        for variable in variables9:
            if variable.get() == 'Very Low':
                variable.set('1')
                variable.set(int(variable.get()))
            elif variable.get() == 'Low':
                variable.set('2')
                variable.set(int(variable.get()))
            elif variable == variables8 and variable.get() == 'Average':
                variable.set('3')
                variable.set(int(variable.get()))
            elif variable.get() == 'Below Average':
                variable.set('3')
                variable.set(int(variable.get()))
            elif variable.get() == 'Average':
                variable.set('4')
                variable.set(int(variable.get()))
            elif variable.get() == 'Above Average':
                variable.set('5')
                variable.set(int(variable.get()))
            elif variable.get() == 'High':
                variable.set('6')
                variable.set(int(variable.get()))
            elif variable.get() == 'Very High':
                variable.set('7')
                variable.set(int(variable.get()))
            elif variable.get() == 'Maximum':
                variable.set('8')
                variable.set(int(variable.get()))
            elif variable.get() == 'Peak':
                variable.set('9')
                variable.set(int(variable.get()))

        neigh2 = KNeighborsClassifier(n_neighbors=2, metric='euclidean')
        neigh2.fit(x, y)

        train = neigh2.predict(
            [[age, gender,
              int(self.airpollution_combobox.get()),
              int(self.alcoholuse_combobox.get()),
              int(self.dustalergy_combobox.get()),
              int(self.occupationalhazard_combobox.get()),
              int(self.geneticrisk_combobox.get()),
              int(self.croniclung_combobox.get()),
              int(self.balanceddiet_combobox.get()),
              int(self.obesity_combobox.get()),
              int(self.smoking_combobox.get()),
              int(self.passivesmoker_combobox.get()),
              int(self.chestpain_combobox.get()),
              int(self.coughingofblood_combobox.get()),
              int(self.fatigue_combobox.get()),
              int(self.weightloss_combobox.get()),
              int(self.shortnessofbreath_combobox.get()),
              int(self.wheezing_combobox.get()),
              int(self.swallowingdiffi_combobox.get()),
              int(self.clubbing_combobox.get()),
              int(self.drycough_combobox.get()),
              int(self.frequentcold_combobox.get()),
              int(self.snoring_combobox.get())]])

        self.hasiltrain = train[0]
        hasil_prediksi = ttk.Label(self.frame, text=self.hasiltrain)
        hasil_prediksi.grid(row=2, column=2)

    # def toggle_mode(self):
    #     if self.mode_switch.instate(["selected"]):
    #         self.style.theme_use("forest-light")
    #     else:
    #         self.style.theme_use("forest-dark")
    def grafik(self):
        self.dfgraf = pd.read_excel('data1 copy.xlsx')

        column_data = self.dfgraf['Level']

        print(column_data)

        nilai_counts = self.dfgraf['Level'].value_counts()
        sales_data = nilai_counts

        plt.rcParams["axes.prop_cycle"] = plt.cycler(
            color=["#4C2A85", "#BE96FF", "#957DAD", "#5E366E", "#A98CCC"])

        fig1, ax1 = plt.subplots()
        ax1.bar(sales_data.keys(), sales_data)
        ax1.set_title("Sales by Product")
        ax1.set_xlabel("Level")
        ax1.set_ylabel("Sales")
        # plt.show()

        # Create a window and add charts
        # root = tk.Tk()
        # root.title("Dashboard")
        # root.state('zoomed')

        side_frame = tk.Frame(self.tab2, bg="#4C2A85")
        side_frame.pack(side="left", fill="both")

        label = tk.Label(side_frame, text="Dashboard",
                         bg="#4C2A85", fg="#FFF", font=25)
        label.pack(pady=50, padx=20)

        charts_frame = tk.Frame(self.tab2)
        charts_frame.pack()

        upper_frame = tk.Frame(charts_frame)
        upper_frame.pack(fill="both")

        canvas1 = FigureCanvasTkAgg(fig1, upper_frame)
        canvas1.draw()
        canvas1.get_tk_widget().pack(fill="both", expand=True)

        lower_frame = tk.Frame(charts_frame)
        lower_frame.pack(fill="both", expand=True)

        # root.mainloop()

    def on_click(self, event):
        item = self.treeview.item(self.treeview.focus())
        popup = tk.Toplevel(root)
        # popup.geometry("200x200")
        label = tk.Label(popup, text=item["text"])
        label.pack()
        popup.update()
        popup.minsize(root.winfo_width(), root.winfo_height())
        x_cordinate = int((root.winfo_screenwidth()/2) -
                          (root.winfo_width()/2))
        y_cordinate = int((root.winfo_screenheight()/2) -
                          (root.winfo_height()/2))
        popup.geometry("+{}+{}".format(x_cordinate, y_cordinate))
        for key in item:
            if key != "text":
                sub_label = tk.Label(popup, text=f"{key}: {item[key]}")
                sub_label.pack()


App(root)
root.mainloop()

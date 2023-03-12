# package yang dibutuhkan
from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import font as tkfont
import pathlib
import time
from geopy.geocoders import Nominatim
import geocoder
import json
from urllib import request
import os.path
import tkinter as tk
import cv2
from PIL import Image, ImageTk
import util
from fpdf import FPDF
from tkinter import ttk


# Window Utama


class App():
    def __init__(self):
        self.window = Tk()
        self.window.geometry("650x450")
        self.window.configure(bg='#1995ad')
        self.window.resizable(False, False)
        self.window.title("Student Presence")

        icon_image = PhotoImage(file='UNESA.png')
        img2 = PhotoImage(file='UNESA.png')
        self.window.iconphoto(False, icon_image)

        font = Font(bold=True)
        border = Border(left=Side(border_style='thin', color='00000000'),
                        right=Side(border_style='thin', color='00000000'),
                        top=Side(border_style='thin', color='00000000'),
                        bottom=Side(border_style='thin', color='00000000'))

        alignment = Alignment(horizontal='center', vertical='center')


# untuk mengecek file backend ada atau tidak
        self.file = pathlib.Path('data.xlsx')
        if self.file.exists():
            pass
        else:
            self.file = Workbook()
            sheet = self.file.active
            sheet['a1'] = 'Hari'
            a1 = sheet['a1']
            a1.font = font
            a1.border = border
            a1.alignment = alignment
            sheet['b1'] = 'Waktu'
            b1 = sheet['b1']
            b1.font = font
            b1.border = border
            b1.alignment = alignment
            sheet['c1'] = 'Nama lengkap'
            c1 = sheet['c1']
            c1.font = font
            c1.border = border
            c1.alignment = alignment
            sheet['d1'] = 'NIM'
            d1 = sheet['d1']
            d1.font = font
            d1.border = border
            d1.alignment = alignment
            sheet['e1'] = 'Akun Unesa'
            e1 = sheet['e1']
            e1.font = font
            e1.border = border
            e1.alignment = alignment
            sheet['f1'] = 'Mata Kuliah'
            f1 = sheet['f1']
            f1.font = font
            f1.border = border
            f1.alignment = alignment
            sheet['g1'] = 'Lokasi '
            g1 = sheet['g1']
            g1.font = font
            g1.border = border
            g1.alignment = alignment
            self.file.save('data.xlsx')

        self.nama_lengkap = StringVar()
        self.nim = StringVar()
        self.umur = StringVar()
        self.matakuliah = StringVar()
        self.akun = StringVar()

# fungsi matakuliah
        y = time.strftime('%A')
        self.x = y.lower()
        if self.x == "monday":
            self.mata_kuliah = ['Kalkulus', 'Praktikum PD', 'Pend Pancasila']
        elif self.x == "tuesday":
            self.mata_kuliah = ['Binggris', 'Teori PD', 'Bindo']
        elif self.x == 'wednesday':
            self.mata_kuliah = ['Arkom', 'Probabilitas', 'Tambahan PD']
        elif self.x == 'thursday':
            self.mata_kuliah = ['Olahraga']
        elif self.x == 'friday':
            self.mata_kuliah = ['Uas Pemrograman Dasar']
        else:
            self.mata_kuliah = ['Kosong']

 # pengaturan font
        self.font_heading = tkfont.Font(family="sans-serif", weight='bold')
        self.font_jam = tkfont.Font(
            family="sans-serif", size=15, weight='bold')
        self.font_body = tkfont.Font(family="sans-serif", weight='bold')
        self.font_entry = tkfont.Font(family="sans-serif")
        self.font_camera = tkfont.Font(
            family="sans-serif", size=25, weight='bold')


# komponen komponen label
        Label(
            self.window, text="Input NIM", font=self.font_body, bg='#a1d6e2', fg='#011a27').place(x=45, y=70)
        Label(
            self.window, text="Name        ", font=self.font_body, bg='#a1d6e2', fg='#011a27').place(x=45, y=120)
        Label(
            self.window, text="Account   ", font=self.font_body, bg='#a1d6e2', fg='#011a27').place(x=45, y=170)
        Label(
            self.window, text="Subject    ", font=self.font_body, bg='#a1d6e2', fg='#011a27').place(x=45, y=220)
        Label(
            self.window, text="Picture     ", font=self.font_body, bg='#a1d6e2', fg='#011a27').place(x=45, y=280)
        img1 = PhotoImage(file='upload.png')
        # f = Frame(self.window, bd=3, bg='black',
        #           width='130', height='130', relief=GROOVE)
        # f.place(x=390, y=200)
        # self.capture_image = Label(
        #     f, image=img1)
        # self.capture_image.place(x=0, y=0)

# komponen komponen entry
        nim_entry = Entry(
            self.window, textvariable=self.nim, width=30, bd=2, font=self.font_entry, bg='#f1f1f2', fg='#011a27')
        nama_lengkap_entry = Entry(
            self.window, textvariable=self.nama_lengkap, width=41, bd=2, font=self.font_entry, bg='#f1f1f2', fg='#011a27')
        akun_entry = Entry(self.window, textvariable=self.akun,
                           width=41, bd=2, font=self.font_entry, bg='#f1f1f2', fg='#011a27')
        self.matakuliah_combobox = Combobox(
            self.window, values=self.mata_kuliah, font=self.font_entry, state='r', width=11, height=2)
        tombol_cari = Button(self.window, text="Find🔍", width=5, height=2, bd=2,
                             bg='#bcbabe', fg='#011a27', command=self.database)
        tombol_submit = Button(self.window, text="Submit", bg='#bcbabe', fg='#011a27',
                               command=self.tombolsubmit, width=15, height=2)
        tombol_clear = Button(self.window, text="Clear", bg='#bcbabe', fg='#011a27',
                              command=self.tombolclear, width=15, height=2)
        tombol_camera = Button(self.window, text="Camera 📷", bg='#bcbabe', fg='#011a27',
                               command=self.mulai1, width=10, height=2)
        tombol_pdf = Button(self.window, text="print", bg='#bcbabe', fg='#011a27',
                            command=self.print_pdf, width=10, height=2)
        tombol_report = Button(self.window, text='report', bg='#bcbabe', fg='#011a27',
                               command=self.load_data, width=10, height=2)
        self.jam()

# letak komponen tombol tombol
        nim_entry.place(x=185, y=70)
        nama_lengkap_entry.place(x=185, y=120)
        akun_entry.place(x=185, y=170)
        self.matakuliah_combobox.place(x=185, y=220)
        self.matakuliah_combobox.set('select')
        tombol_cari.place(x=484, y=63)
        tombol_submit.place(x=185, y=350)
        tombol_clear.place(x=394, y=350)
        tombol_camera.place(x=185, y=270)
        tombol_pdf.place(x=300, y=270)
        tombol_report.place(x=225, y=30)


# mendapatakan adress dengan ip
        h = geocoder.ip('me')
        f = h.ip
        geoloc = Nominatim(user_agent='GetLoc')
        self.url = "https://ipapi.co/"
        self.ip = str(f)
        self.request = request.urlopen(self.url+self.ip+"/json")
        self.data_json = json.loads(self.request.read())
        g = str(self.data_json['latitude'])
        h = str(self.data_json['longitude'])
        self.locname = geoloc.reverse(f'{g},{h}')
        self.locname = str(self.locname)

# database mahasiswa

    def database(self):
        self.akun_unesa = {22031554001: 'varel.22001@mhs.unesa.ac.id',
                           22031554017: 'rizky.22017@mhs.unesa.ac.id',
                           22031554029: 'Hendra.22029@mhs.unesa.ac.id',
                           22031554019: 'azaria.22019@mhs.unesa.ac.id',
                           22031554003: 'fadhilah.22003@mhs.unessa.ac.id',
                           22031554049: 'ilham.22049@mhs.unesa.ac.id',
                           22031554039: 'nurhalizah.22039@mhs.unesa.ac.id',
                           22031554015: 'abdini.22015@mhs.unesa.ac.id',
                           22031554041: 'alivia.22041@mhs.unesa.ac.id',
                           22031554013: 'shafa.22013@mhs.unesa.ac.id ',
                           22031554056: 'waw.5567@mhs.unesa.ac.id',
                           22031554043: 'Riva.22043@mhs.unesa.ac.id ',
                           22031554051: 'nabila.22051@mhs.unesa.ac.id',
                           22031554037: 'devin.22037@mhz.unesa.ac.id',
                           22030224031: 'ahmad.22031@mhs.unesa.ac.id',
                           22040704262: 'denta.22047@mhs.unesa.ac.id',
                           20080324017: 'rofaqazharul.20017@mhs.unesa.ac.id',
                           22031554055: 'michael.22055@mhs.unesa.ac.id',
                           19: 'sembilan belas',
                           20: 'dua puluh'}

        self.fullname = {22031554001: "Varel Dhany Ekamartha ",
                         22031554017: "Rizky Yanuar Kristianto",
                         22031554029: "Hendra Cahyono",
                         22031554019: "Azaria Syahla Fitan Adibah ",
                         22031554003: "Fadhilah Nuria Shinta",
                         22031554049: "Ilham Warmandev",
                         22031554039: "Nur Halizah Amrita",
                         22031554015: "Abdini Qolbi Sahlina",
                         22031554041: "Alivia Nayla Wibisono ",
                         22031554013: "Shafa Nasywa Dhiya",
                         22031554056: "Wawww",
                         22031554043: "Riva Dian Ardiansyah ",
                         22031554051: "Nabila Aulia Arfiani ",
                         22031554037: "Devin Aulia Asshafa",
                         22030224031: "Ahmad Ardiansyah Putra",
                         22040704262: "Denta Lasonda",
                         20080324017: "Rofaq Azharul Ahmad",
                         22031554055: "Michael luwi pallea ",
                         19: "songolas",
                         20: "kalehdoso"}
        angka1 = self.nim.get()
        angka1 = int(angka1)
        self.nama_lengkap.set('')
        self.umur.set('')
        self.akun.set('')
        if angka1 in self.akun_unesa:
            if angka1 in self.fullname:
                self.nama_lengkap.set(self.fullname[angka1])
                self.akun.set(self.akun_unesa[angka1])
        else:
            messagebox.showinfo('info', 'NIM Not Found')

# window camera
    def mulai1(self):
        self.main_window = tk.Toplevel(self.window)
        self.main_window.geometry("1100x520+20+50")
        self.main_window.resizable(False, False)
        self.main_window.configure(bg='#1995ad')

        self.register_new_user_button_main_window = util.get_button(self.main_window, 'cheese😃', 'grey',
                                                                    self.register_new_user, fg='black')
        self.register_new_user_button_main_window.place(x=750, y=210)

        self.webcam_label = util.get_img_label(self.main_window)
        self.webcam_label.place(x=10, y=0, width=700, height=500)

        self.add_webcam(self.webcam_label)

        self.db_dir = './db'
        if not os.path.exists(self.db_dir):
            os.mkdir(self.db_dir)

    def add_webcam(self, label):
        if 'cap' not in self.__dict__:
            self.cap = cv2.VideoCapture(0)

        self._label = label
        self.process_webcam()

    def process_webcam(self):
        ret, frame = self.cap.read()

        self.most_recent_capture_arr = frame
        img_ = cv2.cvtColor(self.most_recent_capture_arr, cv2.COLOR_BGR2RGB)
        self.most_recent_capture_pil = Image.fromarray(img_)
        imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
        self._label.imgtk = imgtk
        self._label.configure(image=imgtk)

        self._label.after(20, self.process_webcam)

    def register_new_user(self):
        self.register_new_user_window = tk.Toplevel(self.window)
        self.register_new_user_window.geometry("1100x520")
        self.register_new_user_window.configure(bg='#1995ad')

        self.accept_button_register_new_user_window = util.get_button(
            self.register_new_user_window, 'Accept', 'grey', self.accept_register_new_user, fg='black')
        self.accept_button_register_new_user_window.place(x=750, y=300)

        self.try_again_button_register_new_user_window = util.get_button(
            self.register_new_user_window, 'Try again', 'grey', self.try_again_register_new_user, fg='black')
        self.try_again_button_register_new_user_window.place(x=750, y=400,)

        self.capture_label = util.get_img_label(self.register_new_user_window)
        self.capture_label.place(x=10, y=0, width=700, height=500)

        self.add_img_to_label(self.capture_label)

        self.text_label_register_new_user = Label(
            self.register_new_user_window, text=f'{self.nama_lengkap.get()}\nAre You Satisfied?',
            bg='#1995ad', fg='#f1f1f2', font=self.font_camera)
        self.text_label_register_new_user.place(x=750, y=200)

    def try_again_register_new_user(self):
        self.register_new_user_window.destroy()

    def add_img_to_label(self, label):
        self.imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
        label.imgtk = self.imgtk
        label.configure(image=self.imgtk)
        self.register_new_user_capture = self.most_recent_capture_arr.copy()

    def start(self):
        self.window.mainloop()

    def accept_register_new_user(self):
        self.name = self.nama_lengkap.get()
        self.acc = self.nim.get()
        self.waktu = time.strftime('%H.%M.%S-%p')
        cv2.imwrite(os.path.join(self.db_dir, '{}_{}.jpg'.format(self.name, self.waktu)),
                    self.register_new_user_capture)

        util.msg_box('Success!', 'User was registered successfully !')

        self.add_image()

        self.register_new_user_window.destroy()
        self.close_camera()

    def close_camera(self):
        self.main_window.destroy()

# fungsi tombol tombol
    def tombolsubmit(self):
        getlokasi = self.locname
        getnama = self.nama_lengkap.get()
        getnim = self.nim.get()
        getakun = self.akun.get()
        getmatakuliah = self.matakuliah_combobox.get()
        self.getjam = time.strftime('%H:%M:%S')
        gethari = time.strftime('%A(%x)')
        self.file = openpyxl.load_workbook('data.xlsx')
        sheet = self.file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=gethari)
        sheet.cell(column=2, row=sheet.max_row, value=self.getjam)
        sheet.cell(column=3, row=sheet.max_row, value=getnama)
        sheet.cell(column=4, row=sheet.max_row, value=getnim)
        sheet.cell(column=5, row=sheet.max_row, value=getakun)
        sheet.cell(column=6, row=sheet.max_row, value=getmatakuliah)
        sheet.cell(column=7, row=sheet.max_row, value=getlokasi)
        self.file.save('data.xlsx')

        messagebox.showinfo('info', 'thankyou')

        self.nama_lengkap.set('')
        self.akun.set('')
        self.nim.set('')
        self.umur.set('')
        self.matakuliah_combobox.set('select')

    def tombolclear(self):
        self.nama_lengkap.set('')
        self.nim.set('')
        self.umur.set('')
        self.akun.set('')
        self.matakuliah_combobox.set('select')

    def jam(self):
        self.widgetjam = Label(self.window, font=self.font_jam, bg='#1995ad',
                               fg='#a1d6e2')
        self.widgetjam.place(x=499, y=20)
        self.widgetanggal = Label(self.window, font=self.font_jam, bg='#1995ad',
                                  fg='#a1d6e2')
        self.widgetanggal.place(x=45, y=20)

        self.x = time.strftime('%H:%M:%S')
        self.widgetjam.config(text=self.x)
        self.widgetjam.after(200, self.jam)

        self.y = time.strftime('%A')
        self.widgetanggal.config(text=self.y)

    def print_pdf(self):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(
            200, 10, txt=f'nama:\t{self.nama_lengkap.get()}', ln=1, align="C")
        pdf.cell(200, 10, txt=f'nim:\t{self.nim.get()}', ln=2, align="C")
        pdf.cell(
            200, 10, txt=f'akun unesa:\t{self.akun.get()}', ln=3, align="C")
        pdf.cell(
            200, 10, txt=f'alamat absen:\t{self.locname}', ln=4, align="C")
        pdf.image(
            f'db\{self.nama_lengkap.get()}_{self.waktu}.jpg', 10, 20, 33, 0)
        pdf.output(f"{self.nama_lengkap.get()}_.pdf")
        messagebox.showinfo('info', 'pdf telah dibuat')

    def load_data(self):
        windowtree = tk.Tk()
        path = 'data.xlsx'
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        list_values = list(sheet.values)
        cols = list_values[0]
        tree = ttk.Treeview(windowtree, column=cols, show="headings")

        # scrollbar = ttk.Scrollbar(
        #     windowtree, orient=tk.VERTICAL, command=tree.yview)
        # tree.configure(yscroll=scrollbar.set)
        # scrollbar.grid(row=0, column=1, sticky='ns')

        for col_name in cols:
            tree.heading(col_name, text=col_name)
        tree.pack(expand=True, fill='y')

        for value_tuple in list_values[1:]:
            tree.insert('', tk.END, values=value_tuple)

    def add_image(self):
        # img = Image.open(self.register_new_user_capture)
        # self.imgtk = ImageTk.PhotoImage(image=self.register_new_user_capture)
        # self.capture_image.configure(image=self.imgtk)
        # resize_image = self.imgtk.resize((130, 130))
        # img2 = ImageTk.PhotoImage(resize_image)
        # self.capture_image.configure(image=img2)
        img_label = tk.Label(self.window)
        img_label.image = tk.PhotoImage(file="upload.png")
        img_label['image'] = img_label.image

        img_label.place(x=390, y=200)


# fungsi untuk looping program
if __name__ == "__main__":
    app = App()
    app.start()

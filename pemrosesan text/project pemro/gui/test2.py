import azure.cognitiveservices.speech as speechsdk
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, Text
from PIL import Image, ImageTk, ImageSequence
import itertools
import joblib
import re
import nltk
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pathlib
import re
from nltk.corpus import stopwords
from sklearn.feature_extraction.text import CountVectorizer
from datetime import datetime
import tkinter.font as tkFont
from tkinter import filedialog
import csv
import pandas as pd
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
from tqdm import tqdm
import time
nltk.download('stopwords')


class App():
    def __init__(self, master):
        self.font = Font(bold=True)
        self.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thin', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'))
        self.alignment = Alignment(horizontal='center', vertical='center')
        self.file = pathlib.Path('data.xlsx')
        if self.file.exists():
            pass
        else:
            self.file = Workbook()
            sheet = self.file.active
            headers = ['Waktu', 'Text', 'Sentimen']
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = self.font
                cell.border = self.border
                cell.alignment = self.alignment
            self.file.save('data.xlsx')
        self.loaded_vectorizer = joblib.load(
            r'C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\tfidf_vectorizer.pkl')
        self.loaded_model = joblib.load(
            r'C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\rbf100_model.pkl')

        self.window = master
        self.window.title("Analisis Sentimen")
        self.window.resizable(True, True)
        self.window.geometry("720x720")
        self.style = ttk.Style(self.window)
        self.icon = ImageTk.PhotoImage(Image.open(
            r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\mic.png"))
        self.window.iconphoto(False, self.icon)
        self.window.tk.call(
            "source", r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\forest-dark.tcl")
        self.window.tk.call(
            "source", r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\forest-light.tcl")
        self.style.theme_use("forest-dark")
        self.notebook = ttk.Notebook(self.window)
        self.tab1 = ttk.Frame(self.notebook)
        self.tab2 = ttk.Frame(self.notebook)

        self.notebook.add(self.tab1, text='Speech to text')
        self.notebook.add(self.tab2, text='Upload file')

        self.notebook.pack(expand=True, fill='both')
        self.frame = ttk.Frame(self.tab1)

        self.frame.place(x=0, y=0, width=720, height=720)
        self.frame2 = ttk.Frame(self.tab2)

        self.frame2.place(x=0, y=0, width=720, height=720)

        self.mode_switch = ttk.Checkbutton(
            self.frame, text="Mode", style="Switch", command=lambda: self.toggle_mode())
        self.mode_switch.place(x=580, y=15)

        if self.mode_switch.instate(["selected"]):
            self.mode = 'default'
        else:
            self.mode = 'dark_background'
        bg1 = '#a9e1b1'
        self.last_mode = "dark"
        self.text_box = tk.Text(self.frame, height=16, width=45, bg="white",
                                borderwidth=0, font=('Arial', 20))

        self.text_box.config(bg=bg1, fg="white",
                             insertbackground="white")
        self.text_box.place(x=16, y=50)
        separator = ttk.Separator(self.frame2)
        separator.place(x=0, y=65, width=720, height=2)
        self.sentimen_button = ttk.Button(self.frame, text="Analis Sentimen",
                                          style="Accent.TButton", command=self.analyze_sentiment_speech)

        self.sentimen_button.place(x=280, y=600, width=150, height=50)
        font_style = tkFont.Font(family="Arial", size=25)
        label = ttk.Label(self.frame, text='Speech to Text', font=font_style)
        label.place(x=10, y=5)
        gif_trash_path = r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\trash-can-unscreen.gif"
        gif_trash = Image.open(gif_trash_path)
        sequence_trash = itertools.cycle(ImageTk.PhotoImage(img.resize((100, 100)))
                                         for img in ImageSequence.Iterator(gif_trash))
        image_trash = next(sequence_trash)
        self.clear_button = tk.Button(
            self.frame, command=self.clear_text, image=image_trash, borderwidth=0, highlightthickness=0, bg=bg1)
        self.clear_button.place(x=400, y=415)
        self.update_button_image(self.clear_button, sequence_trash, 33)
        gif_mic_path = r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\microphone-unscreen.gif"
        gif_mic = Image.open(gif_mic_path)
        sequence_mic = itertools.cycle(ImageTk.PhotoImage(img.resize((100, 100)))
                                       for img in ImageSequence.Iterator(gif_mic))
        image_mic = next(sequence_mic)
        self.start_button = tk.Button(
            self.frame, command=self.recognize_speech, image=image_mic, borderwidth=0, highlightthickness=0, bg=bg1)
        self.start_button.place(x=200, y=415)
        self.update_button_image(self.start_button, sequence_mic, 33)
        gif_upload = r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\upload2-unscreen.gif"
        gif_upload = Image.open(gif_upload)
        sequence_upload = itertools.cycle(ImageTk.PhotoImage(img.resize((50, 50)))
                                          for img in ImageSequence.Iterator(gif_upload))
        image_upload = next(sequence_upload)
        self.load_button = tk.Button(
            self.frame2, command=self.load_data, image=image_upload, borderwidth=0, highlightthickness=0, bg='#313131')
        self.load_button.place(x=500, y=10)
        self.update_button_image(self.load_button, sequence_upload, 33)
        self.alamat_file = ttk.Entry(self.frame2)
        self.alamat_file.insert(0, ".....")
        self.alamat_file.bind(
            "<FocusIn>", lambda e: self.alamat_file.delete('0', 'end'))
        self.alamat_file.place(x=120, y=23, width=370, height=30)
        cols = ("columns1", 'columns2', "columns3")
        treeScroll = ttk.Scrollbar(self.frame2)
        treeScroll.pack(side="right", fill="y")
        self.treeview = ttk.Treeview(self.frame2, show="headings",
                                     yscrollcommand=treeScroll.set, columns=cols, height=13)
        self.treeview.place(x=26, y=80, width=650, height=550)
        self.treeview.heading('#1', text='columns1')
        self.treeview.heading('#2', text='columns2')
        self.treeview.heading('#3', text='columns3')
        self.treeview.column('columns1', width=50)
        self.treeview.column('columns2', width=70)
        self.treeview.column('columns3', width=50)

    def load_data(self):
        # Meminta pengguna untuk memilih file Excel atau CSV
        self.path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")])
        if not self.path:  # Jika pengguna membatalkan dialog, keluar dari fungsi
            return
        self.alamat_file.delete('0', 'end')  # Clear the entry widget
        self.alamat_file.insert('0', self.path)

        # Menghapus entri yang ada di Treeview
        for item in self.treeview.get_children():
            self.treeview.delete(item)

        # Memeriksa ekstensi file dan memuat data sesuai
        if self.path.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(self.path)
            sheet = workbook.active
            list_values = list(sheet.values)
        elif self.path.endswith('.csv'):
            with open(self.path, newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                list_values = list(reader)

        cols = list_values[0]

        # Mengatur kolom dan heading Treeview
        self.treeview.config(columns=cols)
        for col_name in cols:
            self.treeview.heading(col_name, text=col_name)
            # Atur lebar kolom jika diperlukan
            self.treeview.column(col_name, width=100)
            self.treeview.heading(
                col_name, command=lambda c=col_name: self.on_heading_click(c))

        # Memasukkan data baru ke dalam Treeview
        for value_tuple in list_values[1:]:
            self.treeview.insert('', tk.END, values=value_tuple)

        # Menambahkan scrollbar vertikal
        tree_scroll_y = ttk.Scrollbar(
            self.frame2, orient='vertical', command=self.treeview.yview)
        # Sesuaikan posisi scrollbar sesuai dengan Treeview
        tree_scroll_y.place(x=678, y=80, height=550)

        # Menambahkan scrollbar horizontal
        tree_scroll_x = ttk.Scrollbar(
            self.frame2, orient='horizontal', command=self.treeview.xview)
        # Sesuaikan posisi scrollbar sesuai dengan Treeview
        tree_scroll_x.place(x=32, y=633, width=645)

        # Mengonfigurasi Treeview untuk menggunakan scrollbar
        self.treeview.configure(
            yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

    def update_button_image(self, button, sequence, delay):
        frame = next(sequence)
        button.config(image=frame)
        button.image = frame  # Menyimpan referensi gambar
        root.after(delay, self.update_button_image, button, sequence, delay)

    def clear_text(self):
        self.text_box.delete("1.0", tk.END)

    def toggle_mode(self):
        if self.mode_switch.instate(["selected"]):
            self.style.theme_use("forest-light")
            self.mode_switch.configure(text="Light Mode")
            self.last_mode = "light"
            if hasattr(self, 'button'):
                self.button.configure(bg='#FFFFFF')
            self.load_button.configure(bg='#FFFFFF')
        else:
            self.style.theme_use("forest-dark")
            self.mode_switch.configure(text="Dark Mode")
            self.last_mode = "dark"
            if hasattr(self, 'button'):
                self.button.configure(bg='#313131')
            self.load_button.configure(bg='#313131')

    def insert_data(self):
        self.now = datetime.now()
        self.tanggal = self.now.strftime("%d %B %Y")
        self.text = self.text_box.get("1.0", tk.END)
        self.sentiment = self.sentiment_result
        path = self.file
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_values = [self.tanggal, self.text, self.sentiment]
        sheet.append(row_values)
        workbook.save(path)

    def on_heading_click(self, col):
        self.progress = ttk.Progressbar(
            self.frame2, length=100, mode='determinate')
        # Mendapatkan semua item di Treeview
        all_items = self.treeview.get_children()
        # Membuat DataFrame untuk menyimpan data dan hasil analisis sentimen
        self.df = pd.DataFrame(
            columns=self.treeview['columns'] + ('Sentimen',))

        for item in all_items:
            item_data = self.treeview.item(item)['values']
            column_index = self.treeview['columns'].index(col)
            text = item_data[column_index]
            sentiment = self.analyze_sentiment_file(
                text)  # Melakukan analisis sentimen
            self.df.loc[len(self.df)] = item_data + [sentiment]

        # Menambahkan kolom "Sentimen" ke Treeview jika belum ada
        if 'Sentimen' not in self.treeview['columns']:
            self.treeview['columns'] = self.treeview['columns'] + ('Sentimen',)
            self.treeview.heading('Sentimen', text='Sentimen')
            self.treeview.column('Sentimen', width=100)

        # Memperbarui Treeview dengan data yang telah diperbarui
        for item in self.treeview.get_children():
            self.treeview.delete(item)
        for row in self.df.itertuples():
            self.treeview.insert('', tk.END, values=row[1:])
        bg_color = '#FFFFFF' if self.last_mode == "light" else '#313131'
        gif_download = r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\download-folder-unscreen.gif"
        gif_download = Image.open(gif_download)
        sequence_download = itertools.cycle(ImageTk.PhotoImage(img.resize((50, 50)))
                                            for img in ImageSequence.Iterator(gif_download))
        image_upload = next(sequence_download)
        self.button = tk.Button(self.frame2, image=image_upload,
                                command=self.save_file, bg=bg_color, borderwidth=0, highlightthickness=0)
        self.button.place(x=0, y=0, width=50, height=50)
        self.update_button_image(self.button, sequence_download, 33)

    def save_file(self):
        save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[
                                                      ("Excel files", "*.xlsx"), ("CSV files", "*.csv")])

        # Check if a file path was provided
        if save_file_path:
            # Save the DataFrame to the chosen file
            if save_file_path.endswith('.xlsx'):
                self.df.to_excel(save_file_path, index=False)
            elif save_file_path.endswith('.csv'):
                self.df.to_csv(save_file_path, index=False, encoding='utf-8')

    def recognize_speech(self):
        # Setel kunci dan wilayah dari sumber daya Azure Speech Anda
        speech_key = "5544cd8ec2fb442c8527470731b0daae"
        region = "southeastasia"

        # Buat objek konfigurasi dengan kunci dan wilayah
        speech_config = speechsdk.SpeechConfig(
            subscription=speech_key, region=region)

        # Atur bahasa pengenalan ke Bahasa Indonesia
        speech_config.speech_recognition_language = "id-ID"

        # Buat pengenal ucapan dengan konfigurasi
        speech_recognizer = speechsdk.SpeechRecognizer(
            speech_config=speech_config)

        # Mulai pengenalan ucapan
        print("Mulai berbicara...")
        self.result = speech_recognizer.recognize_once()

        # Cetak hasil pengenalan ucapan
        if self.result.reason == speechsdk.ResultReason.RecognizedSpeech:
            self.text_box.insert(tk.END, self.result.text)
        elif self.result.reason == speechsdk.ResultReason.NoMatch:
            messagebox.showinfo("Hasil", "Tidak ada ucapan yang terdeteksi.")
        elif self.result.reason == speechsdk.ResultReason.Canceled:
            cancellation_details = self.result.cancellation_details
            messagebox.showerror("Error", "Pengenalan ucapan dibatalkan: {}".format(
                cancellation_details.reason))
            if cancellation_details.reason == speechsdk.CancellationReason.Error:
                messagebox.showerror("Error", "Kesalahan: {}".format(
                    cancellation_details.error_details))

    def analyze_sentiment_file(self, text):
        # Lakukan preprocessing pada teks
        text = re.sub(r'[^a-zA-Z\s]', '', text, re.I | re.A)
        text = text.lower()
        stop_words = set(stopwords.words('indonesian'))
        text = ' '.join([word for word in text.split()
                        if word not in stop_words])
        # Transform the cleaned text to the same format as the training data
        text_to_analyze_transformed = self.loaded_vectorizer.transform([text])
        # Muat model yang telah disimpan
        try:
            self.progress.pack()
            self.progress['maximum'] = 100
            for i in range(100):
                time.sleep(0.01)  # simulate time delay
                self.progress['value'] = i  # increment progress bar
                self.frame2.update_idletasks()  # force redraw of the progress bar
                predicted_sentiment = self.loaded_model.predict(
                    text_to_analyze_transformed)
            # Tentukan kelas berdasarkan hasil prediksi
            if predicted_sentiment[0] == 0:
                self.sentiment_result = "negatif"
            elif predicted_sentiment[0] == 1:
                self.sentiment_result = "positif"
            else:
                self.sentiment_result = "netral"
        except Exception as e:
            # Menampilkan pesan kesalahan jika terjadi masalah saat memuat model atau melakukan prediksi
            messagebox.showerror(
                "Error", "Terjadi kesalahan saat menganalisis sentimen: " + str(e))
            self.sentiment_result = None  # Atur hasil sentimen ke None jika terjadi kesalahan
        return self.sentiment_result

    def analyze_sentiment_speech(self):
        # Lakukan preprocessing pada teks
        text = self.text_box.get("1.0", tk.END)
        text = re.sub(r'[^a-zA-Z\s]', '', text, re.I | re.A)
        text = text.lower()
        stop_words = set(stopwords.words('indonesian'))
        text = ' '.join([word for word in text.split()
                        if word not in stop_words])
        # Assuming you're using Sastrawi for Indonesian stemming
        stemmer = StemmerFactory().create_stemmer()
        text = stemmer.stem(text)
        # Transform the cleaned text to the same format as the training data
        text_to_analyze_transformed = self.loaded_vectorizer.transform([text])
        # Muat model yang telah disimpan
        try:
            predicted_sentiment = self.loaded_model.predict(
                text_to_analyze_transformed)

            # Tentukan kelas berdasarkan hasil prediksi
            if predicted_sentiment[0] == 0:
                self.sentiment_result = "negatif"
            elif predicted_sentiment[0] == 1:
                self.sentiment_result = "positif"
            else:
                self.sentiment_result = "netral"

            messagebox.showinfo(
                "Hasil Analisis Sentimen", "Sentimen teks ini adalah: " + self.sentiment_result)
        except Exception as e:
            messagebox.showerror(
                "Error", "Tidak dapat memuat model: " + str(e))
        self.insert_data()


root = tk.Tk()
app = App(root)
root.mainloop()

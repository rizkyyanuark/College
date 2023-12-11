import azure.cognitiveservices.speech as speechsdk
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, Text
from PIL import Image, ImageTk
import joblib
import re
import nltk
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pathlib
import re
from nltk.corpus import stopwords
from tkinter import Tk, Button, PhotoImage
from sklearn.feature_extraction.text import CountVectorizer
nltk.download('stopwords')


class App():
    def __init__(self, master):
        self.font = Font(bold=True)
        self.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thin', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'))
        self.alignment = Alignment(horizontal='center', vertical='center')
        self.file = pathlib.Path('data.xlsx')
        if self.file.exists():
            pass
        else:
            self.file = Workbook()
            sheet = self.file.active
            headers = ['Hari', 'Waktu', 'text', 'Sentimen']

            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = self.font
                cell.border = self.border
                cell.alignment = self.alignment
            self.file.save('data.xlsx')

        self.window = master
        self.window.title("Analisis Sentimen")
        self.window.resizable(False, False)
        self.window.geometry("720x720")
        self.style = ttk.Style(self.window)
        self.icon = ImageTk.PhotoImage(Image.open(
            r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\mic.png"))
        self.window.iconphoto(False, self.icon)
        self.window.tk.call(
            "source", r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\forest-dark.tcl")
        self.window.tk.call(
            "source", r"C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\forest-light.tcl")
        self.style.theme_use("forest-dark")
        self.notebook = ttk.Notebook(self.window)
        self.tab1 = ttk.Frame(self.notebook)
        self.tab2 = ttk.Frame(self.notebook)

        self.notebook.add(self.tab1, text='Speech to text')
        self.notebook.add(self.tab2, text='Upload file')

        self.notebook.pack(expand=True, fill='both')
        self.frame = ttk.Frame(self.tab1)
        # Menentukan ukuran frame
        self.frame.place(x=0, y=0, width=720, height=720)

        self.text_box = tk.Text(self.frame, height=18, width=45, bg="white",
                                borderwidth=0, font=('Arial', 20))
        # Mengatur warna latar belakang dan warna teks untuk konsistensi dengan ttk
        self.text_box.config(bg="#1e2931", fg="white",
                             insertbackground="white")
        self.text_box.place(x=16, y=20)

        self.sentimen_button = ttk.Button(self.frame, text="Analis Sentimen",
                                          style="Accent.TButton", command=self.analyze_sentiment)
        # Menentukan posisi button di dalam frame
        self.sentimen_button.place(x=400, y=615, width=150, height=50)
        # Menambahkan tombol hapus
        with Image.open(
                r'C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\trash.png') as img:
            mic_resize = img.resize((50, 50))
            self.image_trash = ImageTk.PhotoImage(mic_resize)
        self.clear_button = tk.Button(self.frame, command=self.clear_text, image=self.image_trash,
                                      width=50, height=50, borderwidth=0, highlightthickness=0)
        self.clear_button.place(x=200, y=615)

        # Ganti dengan path yang benar ke file gambar Anda
        with Image.open(r'C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\2176428.png') as img:
            mic_resize = img.resize((50, 50))
            self.image_mic = ImageTk.PhotoImage(mic_resize)

        # Membuat tombol dengan gambar
        self.start_button = Button(self.frame, command=self.recognize_speech, image=self.image_mic,
                                   width=50, height=50, borderwidth=0, highlightthickness=0)
        self.start_button.place(x=100, y=615)

    # Fungsi untuk menghapus teks dari kotak teks

    def clear_text(self):
        self.text_box.delete("1.0", tk.END)

    # Fungsi pengenalan ucapan

    def recognize_speech(self):
        # Setel kunci dan wilayah dari sumber daya Azure Speech Anda
        speech_key = "5544cd8ec2fb442c8527470731b0daae"
        region = "southeastasia"

        # Buat objek konfigurasi dengan kunci dan wilayah
        speech_config = speechsdk.SpeechConfig(
            subscription=speech_key, region=region)

        # Atur bahasa pengenalan ke Bahasa Indonesia
        speech_config.speech_recognition_language = "id-ID"

        # Buat pengenal ucapan dengan konfigurasi
        speech_recognizer = speechsdk.SpeechRecognizer(
            speech_config=speech_config)

        # Mulai pengenalan ucapan
        print("Mulai berbicara...")
        self.result = speech_recognizer.recognize_once()

        # Cetak hasil pengenalan ucapan
        if self.result.reason == speechsdk.ResultReason.RecognizedSpeech:
            self.text_box.insert(tk.END, self.result.text)
        elif self.result.reason == speechsdk.ResultReason.NoMatch:
            messagebox.showinfo("Hasil", "Tidak ada ucapan yang terdeteksi.")
        elif self.result.reason == speechsdk.ResultReason.Canceled:
            cancellation_details = self.result.cancellation_details
            messagebox.showerror("Error", "Pengenalan ucapan dibatalkan: {}".format(
                cancellation_details.reason))
            if cancellation_details.reason == speechsdk.CancellationReason.Error:
                messagebox.showerror("Error", "Kesalahan: {}".format(
                    cancellation_details.error_details))
        self.analyze_sentiment()

    def analyze_sentiment(self):
        loaded_vectorizer = joblib.load(
            r'C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\bow_vectorizer.pkl')
        # Lakukan preprocessing pada teks
        text = self.text_box.get("1.0", tk.END)
        text = re.sub(r'[^a-zA-Z\s]', '', text, re.I | re.A)
        text = text.lower()
        stop_words = set(stopwords.words('indonesian'))
        text = ' '.join([word for word in text.split()
                        if word not in stop_words])

        # Transform the cleaned text to the same format as the training data
        text_to_analyze_transformed = loaded_vectorizer.transform([text])

        # Muat model yang telah disimpan
        try:
            loaded_model = joblib.load(
                r'C:\Users\rizky\OneDrive\Dokumen\GitHub\test\testpython\pemrosesan text\project pemro\gui\linear1000_model.pkl')
            predicted_sentiment = loaded_model.predict(
                text_to_analyze_transformed)

            # Tentukan kelas berdasarkan hasil prediksi
            if predicted_sentiment[0] == 0:
                sentiment_result = "negatif"
            elif predicted_sentiment[0] == 1:
                sentiment_result = "positif"
            else:
                sentiment_result = "netral"

            messagebox.showinfo(
                "Hasil Analisis Sentimen", "Sentimen teks ini adalah: " + sentiment_result)
        except Exception as e:
            messagebox.showerror(
                "Error", "Tidak dapat memuat model: " + str(e))


root = tk.Tk()
app = App(root)
root.mainloop()
